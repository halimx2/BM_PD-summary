import streamlit as st
import pandas as pd
import io
import os
from openpyxl import load_workbook
from report_extractor import parse_chat_text, extract_report_data

# Streamlit 환경 설정
st.set_page_config(page_title="채팅 리포트 추출기", layout="wide")
st.title("📋 채팅 리포트 추출기 웹 앱")

# 템플릿 엑셀 파일 경로 (프로젝트 루트에 template.xlsx 파일을 위치시키세요)
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'template.xlsx')

# 업로드 위젯: 채팅 txt 파일만 업로드
txt_files = st.file_uploader(
    "채팅 txt 파일 업로드 (여러 개 가능)",
    type="txt",
    accept_multiple_files=True
)

if txt_files:
    # 1) Raw Data 생성
    all_messages = []
    for f in txt_files:
        try:
            text = f.read().decode('utf-8')
        except Exception:
            text = f.read().decode('cp949', errors='ignore')
        all_messages.extend(parse_chat_text(text))
    df_raw = extract_report_data(all_messages)

    if df_raw.empty:
        st.warning("유효한 리포트 데이터가 없습니다.")
    else:
        st.success(f"Raw Data {len(df_raw)}건 추출 완료")
        st.dataframe(df_raw)

        # 2) 로컬 템플릿 로드 및 RawData 시트 덮어쓰기
        try:
            wb = load_workbook(TEMPLATE_PATH)
        except Exception as e:
            st.error(f"템플릿 로드 오류: {e}")
        else:
            sheet_name = 'RawData'
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # 헤더(1행)는 그대로 두고, 2행부터 기존 데이터 삭제
                if ws.max_row > 1:
                    ws.delete_rows(idx=2, amount=ws.max_row - 1)
            else:
                ws = wb.create_sheet(sheet_name, 0)

            # 3) 데이터 쓰기 (2행부터)
            for r, row in enumerate(df_raw.itertuples(index=False, name=None), start=2):
                for c, val in enumerate(row, start=1):
                    ws.cell(row=r, column=c, value=val)

            # 4) 메모리 버퍼에 저장
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # 5) 다운로드 버튼
            st.download_button(
                label="분석 엑셀 다운로드",
                data=output.getvalue(),
                file_name=f"analysis_{pd.Timestamp.now().strftime('%y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
else:
    st.info("채팅 txt 파일을 업로드해주세요.")
